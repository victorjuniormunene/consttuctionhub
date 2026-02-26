from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.template.loader import get_template
from django.db.models import Sum, Count
from apps.orders.models import Order
from apps.accounts.models import CustomUser
from apps.products.models import Product
from apps.suppliers.models import Supplier
from apps.consultations.models import ConsultantApplication
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
import openpyxl
from datetime import datetime, timedelta

@login_required
@user_passes_test(lambda u: u.is_staff)
def reports_dashboard(request):
    """Dashboard showing report options and basic statistics"""
    # Get basic statistics
    total_orders = Order.objects.count()
    total_customers = CustomUser.objects.filter(user_type='customer').count()
    total_suppliers = Supplier.objects.count()
    total_products = Product.objects.count()
    total_consultant_apps = ConsultantApplication.objects.count()

    # Revenue statistics
    total_revenue = Order.objects.filter(status='completed').aggregate(
        total=Sum('total_cost')
    )['total'] or 0

    # Recent orders (last 30 days)
    thirty_days_ago = datetime.now() - timedelta(days=30)
    recent_orders = Order.objects.filter(created_at__gte=thirty_days_ago).count()

    context = {
        'total_orders': total_orders,
        'total_customers': total_customers,
        'total_suppliers': total_suppliers,
        'total_products': total_products,
        'total_consultant_apps': total_consultant_apps,
        'total_revenue': total_revenue,
        'recent_orders': recent_orders,
    }

    return render(request, 'reports/dashboard.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def customer_report_pdf(request):
    """Generate PDF report for customers"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = Paragraph("Customer Report - Construction Hub", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Date
    date_str = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    elements.append(Paragraph(date_str, styles['Normal']))
    elements.append(Spacer(1, 12))

    # Customer data
    customers = CustomUser.objects.filter(user_type='customer').select_related()

    # Customer statistics
    customer_stats = [
        ['Total Customers', len(customers)],
        ['Active Customers', customers.filter(is_active=True).count()],
        ['Inactive Customers', customers.filter(is_active=False).count()],
    ]

    stats_table = Table(customer_stats, colWidths=[200, 100])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 20))

    # Customer details
    customer_data = [['Name', 'Email', 'Phone', 'Location', 'Orders', 'Total Spent']]
    customer_data[0] = [Paragraph(cell, styles['Normal']) for cell in customer_data[0]]

    for customer in customers:
        orders_count = Order.objects.filter(customer=customer).count()
        total_spent = Order.objects.filter(customer=customer, status='completed').aggregate(
            total=Sum('total_cost')
        )['total'] or 0

        row = [
            customer.get_full_name() or customer.username,
            customer.email,
            getattr(customer, 'phone', 'N/A'),
            getattr(customer, 'location', 'N/A'),
            str(orders_count),
            f"KSH {total_spent:,.2f}"
        ]
        customer_data.append(row)

    customer_table = Table(customer_data, colWidths=[80, 120, 80, 80, 50, 80])
    customer_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(customer_table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="customer_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response

@login_required
@user_passes_test(lambda u: u.is_staff)
def supplier_report_pdf(request):
    """Generate PDF report for suppliers"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = Paragraph("Supplier Report - Construction Hub", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Date
    date_str = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    elements.append(Paragraph(date_str, styles['Normal']))
    elements.append(Spacer(1, 12))

    # Supplier data
    suppliers = Supplier.objects.all().select_related('user')

    # Supplier statistics
    supplier_stats = [
        ['Total Suppliers', len(suppliers)],
        ['Active Suppliers', suppliers.filter(user__is_active=True).count()],
        ['Total Products', Product.objects.count()],
    ]

    stats_table = Table(supplier_stats, colWidths=[200, 100])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 20))

    # Supplier details
    supplier_data = [['Company Name', 'Contact Person', 'Email', 'Phone', 'Products', 'Total Sales']]
    supplier_data[0] = [Paragraph(cell, styles['Normal']) for cell in supplier_data[0]]

    for supplier in suppliers:
        products_count = Product.objects.filter(supplier=supplier).count()
        total_sales = Order.objects.filter(
            product__supplier=supplier,
            status='completed'
        ).aggregate(total=Sum('total_cost'))['total'] or 0

        row = [
            supplier.company_name,
            supplier.user.get_full_name() or supplier.user.username,
            supplier.user.email,
            getattr(supplier, 'phone', 'N/A'),
            str(products_count),
            f"KSH {total_sales:,.2f}"
        ]
        supplier_data.append(row)

    supplier_table = Table(supplier_data, colWidths=[100, 80, 120, 80, 50, 80])
    supplier_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(supplier_table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="supplier_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response

@login_required
@user_passes_test(lambda u: u.is_staff)
def customer_report_excel(request):
    """Generate Excel report for customers"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Customer Report'

    # Headers
    headers = ['Name', 'Email', 'Phone', 'Location', 'Orders Count', 'Total Spent', 'Join Date']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Customer data
    customers = CustomUser.objects.filter(user_type='customer').select_related()
    for row_num, customer in enumerate(customers, 2):
        orders_count = Order.objects.filter(customer=customer).count()
        total_spent = Order.objects.filter(customer=customer, status='completed').aggregate(
            total=Sum('total_cost')
        )['total'] or 0

        worksheet.cell(row=row_num, column=1, value=customer.get_full_name() or customer.username)
        worksheet.cell(row=row_num, column=2, value=customer.email)
        worksheet.cell(row=row_num, column=3, value=getattr(customer, 'phone', 'N/A'))
        worksheet.cell(row=row_num, column=4, value=getattr(customer, 'location', 'N/A'))
        worksheet.cell(row=row_num, column=5, value=orders_count)
        worksheet.cell(row=row_num, column=6, value=float(total_spent))
        worksheet.cell(row=row_num, column=7, value=customer.date_joined.strftime('%Y-%m-%d'))

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="customer_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    return response

@login_required
@user_passes_test(lambda u: u.is_staff)
def supplier_report_excel(request):
    """Generate Excel report for suppliers"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Supplier Report'

    # Headers
    headers = ['Company Name', 'Contact Person', 'Email', 'Phone', 'Products Count', 'Total Sales', 'Join Date']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Supplier data
    suppliers = Supplier.objects.all().select_related('user')
    for row_num, supplier in enumerate(suppliers, 2):
        products_count = Product.objects.filter(supplier=supplier).count()
        total_sales = Order.objects.filter(
            product__supplier=supplier,
            status='completed'
        ).aggregate(total=Sum('total_cost'))['total'] or 0

        worksheet.cell(row=row_num, column=1, value=supplier.company_name)
        worksheet.cell(row=row_num, column=2, value=supplier.user.get_full_name() or supplier.user.username)
        worksheet.cell(row=row_num, column=3, value=supplier.user.email)
        worksheet.cell(row=row_num, column=4, value=getattr(supplier, 'phone', 'N/A'))
        worksheet.cell(row=row_num, column=5, value=products_count)
        worksheet.cell(row=row_num, column=6, value=float(total_sales))
        worksheet.cell(row=row_num, column=7, value=supplier.user.date_joined.strftime('%Y-%m-%d'))

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="supplier_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    return response
